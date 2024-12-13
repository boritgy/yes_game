import streamlit as st
import openpyxl

workbook = openpyxl.load_workbook('database.xlsx')
sh = workbook.active

def next_question(team, answer):
    for row in range(1, sh.max_row + 1):
        if sh.cell(row=row, column=1).value == team and answer in str(sh.cell(row=row, column=2).value).split(","):
            st.session_state["default"] = sh.cell(row=row, column=3).value
            st.session_state["format"] = sh.cell(row=row, column=5).value
            st.session_state.answer = ""
            st.session_state["wrong"] = ""
            break
        else:
            st.session_state["wrong"] = "Wrong answer :("
    
if "default" not in st.session_state:
    st.session_state["default"] = ""
    st.session_state["format"] = ""
    st.session_state["wrong"] = ""

st.title("Find the restaurant!")

my_area = st.markdown(st.session_state["default"])

team = st.text_input("The name of your team:", placeholder="Team")
answer = st.text_input("Answer (write 'start' if you are starting out):", placeholder="Answer", key="answer")


st.text("Answer format: " + str(st.session_state['format']))
st.text(str(st.session_state['wrong']))

st.button('Go!', on_click=next_question, args=(team.lower(),answer.lower()))

st.text('In case you are lost, you can call us/message us/send your location on Whatsapp:')
st.text('Bori: +36 30 648 0643')
st.text('Edina: +36 20 272 0464')
st.text('Szabi: +36 30 556 6720')

